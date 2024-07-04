import os
import requests
from io import BytesIO
from PIL import Image
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import tensorflow as tf
from tensorflow.keras.preprocessing.image import ImageDataGenerator # type: ignore
from tensorflow.keras.applications import MobileNetV2 # type: ignore
from tensorflow.keras.layers import Dense, GlobalAveragePooling2D # type: ignore
from tensorflow.keras.models import Model # type: ignore
from tensorflow.keras.optimizers import Adam # type: ignore

# Function to create and compile the model
def create_model(num_classes):
    base_model = MobileNetV2(weights='imagenet', include_top=False, input_shape=(224, 224, 3))
    x = base_model.output
    x = GlobalAveragePooling2D()(x)
    x = Dense(1024, activation='relu')(x)
    predictions = Dense(num_classes, activation='softmax')(x)
    model = Model(inputs=base_model.input, outputs=predictions)

    for layer in base_model.layers:
        layer.trainable = False

    model.compile(optimizer=Adam(learning_rate=0.0001), loss='categorical_crossentropy', metrics=['accuracy'])
    return model

# Function to train the model
def train_model(data_dir, model_path, batch_size=8, epochs=10):
    datagen = ImageDataGenerator(validation_split=0.2, rescale=1./255)
    train_generator = datagen.flow_from_directory(data_dir, target_size=(224, 224), batch_size=batch_size, class_mode='categorical', subset='training')
    validation_generator = datagen.flow_from_directory(data_dir, target_size=(224, 224), batch_size=batch_size, class_mode='categorical', subset='validation')

    model = create_model(num_classes=len(train_generator.class_indices))
    model.fit(train_generator, validation_data=validation_generator, epochs=epochs)

    # Ensure the directory exists before saving the model
    os.makedirs(os.path.dirname(model_path), exist_ok=True)
    model.save(model_path)
    return model, train_generator.class_indices

# Function to fetch image from URL
def fetch_image_from_url(url):
    response = requests.get(url)
    if response.status_code == 200:
        img = Image.open(BytesIO(response.content)).convert('RGB').resize((224, 224))
        return np.array(img)
    else:
        raise Exception(f"Failed to fetch image from URL: {url}, Status Code: {response.status_code}")

# Function to update Excel sheet
def update_excel(sheet_path, data):
    try:
        # Load existing data or create a new workbook if the file does not exist
        if os.path.exists(sheet_path):
            wb = load_workbook(sheet_path)
        else:
            wb = Workbook()

        # Check if the sheet already exists, create new if not
        sheet_name = 'Sheet1'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            ws.title = sheet_name

        # Convert data to DataFrame
        df_new = pd.DataFrame(data)

        # Append new data to the existing sheet
        for r in dataframe_to_rows(df_new, index=False, header=False):
            ws.append(r)

        # Save the workbook
        wb.save(sheet_path)

        print(f"Updated Excel file: {sheet_path}")

    except Exception as e:
        print(f"Error updating Excel file: {sheet_path}, Error: {str(e)}")

# Function to load a trained model
def load_model(model_path):
    return tf.keras.models.load_model(model_path)

# Function to predict objects in an image
def predict_objects(model, class_indices, image):
    image = image / 255.0
    image = np.expand_dims(image, axis=0)
    preds = model.predict(image)
    class_id = np.argmax(preds, axis=1)[0]
    class_name = list(class_indices.keys())[list(class_indices.values()).index(class_id)]
    confidence = preds[0][class_id]
    return class_name, confidence

# Main function to process image, detect objects, and update Excel
def process_image(url):
    try:
        data_dir = r'C:\Users\sanjai\Desktop\nstore qc automation web application django\nstores_QC\nstores_QC\dataset'  # Path to your dataset
        model_dir = r'C:\Users\sanjai\Desktop\nstore qc automation web application django\nstores_QC\nstores_QC\models'  # Directory to save the model
        model_path = os.path.join(model_dir, 'model.h5')
        sheet_path = './results.xlsx'  # Path to your Excel file
        serial_number = 1
        if not os.path.exists(model_path):
            print("Model not found, training the model...")
            model, class_indices = train_model(data_dir, model_path)
        else:
            # Load class indices from training phase
            class_indices = {'class1': 0, 'class2': 1}

        image = fetch_image_from_url(url)
        model = load_model(model_path)
        class_name, confidence = predict_objects(model, class_indices, image)

        result = {
            'URL': url,
            'Serial Number': serial_number,
            'Result': class_name,
            'Confidence': confidence
        }
        print(f"Processed image from URL: {url}, Result: {result['Result']}, Confidence: {result['Confidence']}")
        return result['Result']
    except Exception as e:
        print(f"Error processing image from URL: {url}, Error: {str(e)}")

# Example usage
# if __name__ == "__main__":
    
    # url = 'https://ondc-marketplace.s3.ap-south-1.amazonaws.com/images/dba17b17-c602-4bcd-961c-32f82cd2e146.jpg'  # Replace with a valid URL

    # Ensure the model is trained and saved before processing images
      # Update with your actual class indices

    # Process image and update Excel
